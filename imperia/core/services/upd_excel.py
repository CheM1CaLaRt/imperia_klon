"""
Сервис для заполнения УПД по шаблону Excel
"""
import os
from copy import copy
from datetime import date
from decimal import Decimal
from typing import List, Dict, Optional

import openpyxl
from openpyxl import utils
from openpyxl.styles import Font, Alignment, Border, Side
from django.conf import settings


# Путь к шаблону УПД
# Шаблон должен быть скачан с https://spmag.ru/download/file/4430 или https://spmag.ru/download/file/7487
# и размещен в корне проекта imperia/ как Blank-UPD.xlsx
TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    "Blank-UPD.xlsx"
)


def fill_upd_excel(
    seller_name: str,
    seller_full_name: str,
    seller_address: str,
    seller_inn: str,
    seller_kpp: str = "",
    buyer_name: str = "",
    buyer_full_name: str = "",
    buyer_address: str = "",
    buyer_inn: str = "",
    buyer_kpp: str = "",
    doc_number: str = "",
    doc_date: date = None,
    items: List[Dict] = None,
    currency_name_code: str = "Российский рубль, 643",
    output_path: Optional[str] = None,
) -> bytes:
    """
    Заполняет шаблон УПД данными и возвращает байты файла
    
    Args:
        seller_name: Краткое наименование продавца
        seller_full_name: Полное наименование продавца
        seller_address: Адрес продавца
        seller_inn: ИНН продавца
        seller_kpp: КПП продавца
        buyer_name: Краткое наименование покупателя
        buyer_full_name: Полное наименование покупателя
        buyer_address: Адрес покупателя
        buyer_inn: ИНН покупателя
        buyer_kpp: КПП покупателя
        doc_number: Номер документа
        doc_date: Дата документа
        items: Список товаров/услуг
        currency_name_code: Валюта (по умолчанию "Российский рубль, 643")
        output_path: Путь для сохранения файла (опционально)
    
    Returns:
        bytes: Байты Excel файла
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Шаблон УПД не найден: {TEMPLATE_PATH}")
    
    if doc_date is None:
        doc_date = date.today()
    
    if items is None:
        items = []
    
    # Загружаем шаблон
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # Ищем лист с УПД (может называться "Счет-фактура" или "УПД" или первый лист)
    sheet_names = wb.sheetnames
    ws = None
    for name in ["Счет-фактура", "УПД", "Sheet1"]:
        if name in sheet_names:
            ws = wb[name]
            break
    
    if ws is None:
        ws = wb.active
    
    def safe_set_cell(worksheet, row, col, value):
        """Безопасно устанавливает значение ячейки, обрабатывая объединенные ячейки"""
        try:
            # Проверяем, является ли ячейка частью объединенного диапазона
            # Если да, записываем в верхнюю левую ячейку объединенного диапазона
            for merged_range in list(worksheet.merged_cells.ranges):
                if row >= merged_range.min_row and row <= merged_range.max_row and \
                   col >= merged_range.min_col and col <= merged_range.max_col:
                    # Ячейка в объединенном диапазоне - записываем в верхнюю левую
                    top_left_coord = f"{utils.get_column_letter(merged_range.min_col)}{merged_range.min_row}"
                    try:
                        top_left_cell = worksheet[top_left_coord]
                        top_left_cell.value = value
                        return
                    except:
                        # Пробуем через row/column
                        top_left = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        top_left.value = value
                        return
            
            # Обычная ячейка - устанавливаем значение напрямую
            cell_coordinate = f"{utils.get_column_letter(col)}{row}"
            try:
                cell = worksheet[cell_coordinate]
                cell.value = value
            except:
                # Если не получилось через координаты, пробуем через row/column
                cell = worksheet.cell(row=row, column=col)
                cell.value = value
        except Exception:
            # В крайнем случае просто игнорируем ошибку
            pass
    
    def find_merged_cell_for_column(worksheet, row, col):
        """
        Находит правильную ячейку для записи данных с учетом объединенных ячеек.
        Если колонка объединена, возвращает верхнюю левую ячейку объединенного диапазона.
        Если строка попадает в объединенный диапазон, использует верхнюю строку диапазона.
        """
        # Проверяем все объединенные диапазоны
        for merged_range in list(worksheet.merged_cells.ranges):
            # Проверяем, попадает ли наша ячейка в объединенный диапазон
            if row >= merged_range.min_row and row <= merged_range.max_row and \
               col >= merged_range.min_col and col <= merged_range.max_col:
                # Возвращаем верхнюю левую ячейку объединенного диапазона
                # Используем исходную строку, если она в пределах диапазона
                # Но колонку берем из начала диапазона
                return row, merged_range.min_col
        
        # Если не найдено объединение, возвращаем исходные координаты
        return row, col
    
    # --- Заполнение шапки УПД ---
    # Номер и дата УПД (поле (1))
    # Ищем ячейки с текстом "Счет-фактура", "УПД" или номером документа
    doc_number_set = False
    doc_date_set = False
    
    for row in range(1, 6):
        for col in range(1, 35):
            try:
                cell = ws.cell(row=row, column=col)
                if not cell.value:
                    continue
                    
                val = str(cell.value).strip()
                
                # Ищем заголовок документа
                if not doc_number_set and ("Счет-фактура" in val or ("УПД" in val and "универсальный" not in val.lower())):
                    # Ищем номер документа справа от заголовка
                    for offset in range(1, 10):
                        try:
                            next_cell = ws.cell(row=row, column=col + offset)
                            next_val = str(next_cell.value).strip() if next_cell.value else ""
                            
                            # Если ячейка пустая или содержит только номер, заполняем
                            if not next_val or next_val in ["0", "N", "№"]:
                                safe_set_cell(ws, row, col + offset, doc_number)
                                doc_number_set = True
                                
                                # Ищем дату дальше
                                for date_offset in range(offset + 1, offset + 8):
                                    try:
                                        date_cell = ws.cell(row=row, column=col + date_offset)
                                        prev_cell = ws.cell(row=row, column=col + date_offset - 1)
                                        
                                        prev_val = str(prev_cell.value).strip().lower() if prev_cell.value else ""
                                        date_val = str(date_cell.value).strip() if date_cell.value else ""
                                        
                                        # Если есть "от" перед ячейкой или ячейка пустая
                                        if "от" in prev_val or (not date_val or date_val in ["0", "00.00.0000"]):
                                            safe_set_cell(ws, row, col + date_offset, doc_date.strftime("%d.%m.%Y"))
                                            doc_date_set = True
                                            break
                                    except:
                                        continue
                                break
                        except:
                            continue
                    if doc_number_set:
                        break
            except:
                continue
        if doc_number_set:
            break
    
    # Продавец (2) - ищем метку "Продавец" или "Грузоотправитель"
    seller_found = False
    
    for row in range(1, 25):
        for col in range(1, 15):
            try:
                cell = ws.cell(row=row, column=col)
                if not cell.value:
                    continue
                    
                val = str(cell.value).strip()
                if "Продавец" in val or "Грузоотправитель" in val:
                    seller_row = row
                    seller_col = col
                    
                    # Заполняем наименование (обычно следующая строка)
                    safe_set_cell(ws, seller_row + 1, seller_col, seller_full_name or seller_name)
                    
                    # Заполняем адрес (обычно через строку)
                    safe_set_cell(ws, seller_row + 2, seller_col, seller_address)
                    
                    # ИНН/КПП (может быть через 1-2 строки)
                    inn_kpp = seller_inn
                    if seller_kpp:
                        inn_kpp += f" / {seller_kpp}"
                    
                    # Пробуем несколько строк ниже
                    for offset in [3, 4, 5]:
                        try:
                            test_cell = ws.cell(row=seller_row + offset, column=seller_col)
                            if not test_cell.value or str(test_cell.value).strip() in ["", "0"]:
                                safe_set_cell(ws, seller_row + offset, seller_col, inn_kpp)
                                break
                        except:
                            continue
                    
                    seller_found = True
                    break
            except:
                continue
        if seller_found:
            break
    
    # Покупатель (6) - ищем метку "Покупатель" или "Грузополучатель"
    buyer_found = False
    
    for row in range(1, 35):
        for col in range(1, 15):
            try:
                cell = ws.cell(row=row, column=col)
                if not cell.value:
                    continue
                    
                val = str(cell.value).strip()
                if "Покупатель" in val or "Грузополучатель" in val:
                    buyer_row = row
                    buyer_col = col
                    
                    # Заполняем наименование
                    safe_set_cell(ws, buyer_row + 1, buyer_col, buyer_full_name or buyer_name)
                    
                    # Заполняем адрес
                    safe_set_cell(ws, buyer_row + 2, buyer_col, buyer_address)
                    
                    # ИНН/КПП
                    buyer_inn_kpp = buyer_inn
                    if buyer_kpp:
                        buyer_inn_kpp += f" / {buyer_kpp}"
                    
                    # Пробуем несколько строк ниже
                    for offset in [3, 4, 5]:
                        try:
                            test_cell = ws.cell(row=buyer_row + offset, column=buyer_col)
                            if not test_cell.value or str(test_cell.value).strip() in ["", "0"]:
                                safe_set_cell(ws, buyer_row + offset, buyer_col, buyer_inn_kpp)
                                break
                        except:
                            continue
                    
                    buyer_found = True
                    break
            except:
                continue
        if buyer_found:
            break
    
    # Валюта (7) - согласно стандарту: H15
    currency_found = False
    
    # Пробуем стандартные координаты
    try:
        if ws["H15"].value is None or str(ws["H15"].value).strip() == "":
            safe_set_cell(ws, 15, 8, currency_name_code)  # H15
            currency_found = True
        elif "Валюта" in str(ws["G15"].value or "") or "Денежная" in str(ws["G15"].value or ""):
            safe_set_cell(ws, 15, 8, currency_name_code)
            currency_found = True
    except:
        pass
    
    # Динамический поиск
    if not currency_found:
        for row in range(1, 35):
            for col in range(1, 15):
                try:
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        val = str(cell.value).strip()
                        if "Валюта" in val or "Денежная единица" in val:
                            safe_set_cell(ws, row, col + 1, currency_name_code)
                            currency_found = True
                            break
                except:
                    continue
            if currency_found:
                break
    
    # --- Заполнение таблицы товаров ---
    # Ищем начало таблицы товаров (обычно строка с заголовками)
    table_start_row = None
    header_keywords = ["Наименование", "Единица", "Количество", "Цена", "Стоимость", "НДС"]
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=15, max_row=50), start=15):
        row_values = [str(cell.value or "").lower() for cell in row if cell.value]
        if any(keyword.lower() in " ".join(row_values) for keyword in header_keywords):
            table_start_row = row_idx + 1  # Следующая строка после заголовков
            break
    
    if table_start_row is None:
        # Если не нашли заголовки, начинаем с 24 строки (как в примере)
        table_start_row = 24
    
    total_without_vat = Decimal("0")
    total_vat = Decimal("0")
    total_with_vat = Decimal("0")
    
    # Определяем колонки по заголовкам (если найдены)
    col_num = None  # N п/п
    col_name = None  # Наименование
    col_unit_code = None  # Код ед. изм.
    col_unit = None  # Единица измерения
    col_qty = None  # Количество
    col_price = None  # Цена
    col_total_no_vat = None  # Стоимость без НДС
    col_vat_rate = None  # Ставка НДС
    col_vat_amount = None  # Сумма НДС
    col_total_with_vat = None  # Стоимость с НДС
    
    # Пытаемся определить колонки по заголовкам
    # Проверяем строку заголовков и строку выше (на случай многострочных заголовков)
    for check_row in [table_start_row - 1, table_start_row - 2]:
        if check_row < 1:
            continue
        try:
            header_row = ws[check_row]
            for idx, cell in enumerate(header_row, start=1):
                if not cell.value:
                    continue
                
                # Проверяем, является ли ячейка частью объединенного диапазона
                # Если да, используем колонку объединенного диапазона
                actual_row, actual_col = find_merged_cell_for_column(ws, check_row, idx)
                
                cell_value = str(cell.value).lower().strip()
                # Номер п/п
                if col_num is None and ("п/п" in cell_value or "номер" in cell_value or "№" in cell_value):
                    col_num = actual_col
                # Наименование
                elif col_name is None and ("наименование" in cell_value or "товар" in cell_value or "наимен" in cell_value):
                    col_name = actual_col
                # Единица измерения (не код!)
                elif col_unit is None and "единица" in cell_value and "код" not in cell_value:
                    col_unit = actual_col
                # Количество
                elif col_qty is None and "количество" in cell_value:
                    col_qty = actual_col
                # Цена
                elif col_price is None and "цена" in cell_value and "стоимость" not in cell_value:
                    col_price = actual_col
                # Стоимость без НДС
                elif col_total_no_vat is None and ("без ндс" in cell_value or ("стоимость" in cell_value and "без" in cell_value)):
                    col_total_no_vat = actual_col
                # Ставка НДС
                elif col_vat_rate is None and "ставка" in cell_value and "ндс" in cell_value:
                    col_vat_rate = actual_col
                # Сумма НДС
                elif col_vat_amount is None and "сумма" in cell_value and "ндс" in cell_value:
                    col_vat_amount = actual_col
                # Стоимость с НДС
                elif col_total_with_vat is None and ("с ндс" in cell_value or ("всего" in cell_value and "стоимость" in cell_value)):
                    col_total_with_vat = actual_col
        except:
            continue
    
    # Если не определили колонки, используем значения по умолчанию из примера
    if col_num is None:
        col_num = 2  # B
    if col_name is None:
        col_name = 5  # E
    if col_unit is None:
        col_unit = 14  # N
    if col_qty is None:
        col_qty = 18  # R
    if col_price is None:
        col_price = 23  # W
    if col_total_no_vat is None:
        col_total_no_vat = 27  # AA
    if col_vat_rate is None:
        col_vat_rate = 37  # AK
    if col_vat_amount is None:
        col_vat_amount = 41  # AO
    if col_total_with_vat is None:
        col_total_with_vat = 45  # AS
    
    # Заполняем строки товаров
    current_row = table_start_row
    for idx, item in enumerate(items, start=1):
        name = item.get("name", item.get("title", ""))
        unit = item.get("unit", "шт")
        qty = float(item.get("qty", item.get("quantity", 0)))
        price = float(item.get("price", 0))
        vat_rate_str = item.get("vat_rate", "20%")
        
        amount_without_vat = Decimal(str(qty * price))
        
        # Парсим ставку НДС
        if isinstance(vat_rate_str, str) and vat_rate_str.endswith("%"):
            vat_percent = Decimal(vat_rate_str.strip("%")) / Decimal("100")
        elif isinstance(vat_rate_str, (int, float, Decimal)):
            vat_percent = Decimal(str(vat_rate_str)) / Decimal("100")
        else:
            vat_percent = Decimal("0.20")  # По умолчанию 20%
        
        vat_amount = amount_without_vat * vat_percent
        amount_with_vat = amount_without_vat + vat_amount
        
        total_without_vat += amount_without_vat
        total_vat += vat_amount
        total_with_vat += amount_with_vat
        
        # Заполняем ячейки (используем safe_set_cell для обработки объединенных ячеек)
        # Для каждой колонки находим правильную ячейку с учетом объединенных диапазонов
        # Номер п/п
        if col_num:
            actual_row, actual_col = find_merged_cell_for_column(ws, current_row, col_num)
            safe_set_cell(ws, actual_row, actual_col, idx)
        
        # Наименование товара
        if col_name:
            actual_row, actual_col = find_merged_cell_for_column(ws, current_row, col_name)
            safe_set_cell(ws, actual_row, actual_col, name)
        
        # Единица измерения
        if col_unit:
            actual_row, actual_col = find_merged_cell_for_column(ws, current_row, col_unit)
            safe_set_cell(ws, actual_row, actual_col, unit)
        
        # Количество
        if col_qty:
            actual_row, actual_col = find_merged_cell_for_column(ws, current_row, col_qty)
            safe_set_cell(ws, actual_row, actual_col, qty)
        
        # Цена за единицу
        if col_price:
            actual_row, actual_col = find_merged_cell_for_column(ws, current_row, col_price)
            safe_set_cell(ws, actual_row, actual_col, price)
        
        # Стоимость без НДС
        if col_total_no_vat:
            actual_row, actual_col = find_merged_cell_for_column(ws, current_row, col_total_no_vat)
            safe_set_cell(ws, actual_row, actual_col, float(amount_without_vat))
        
        # Ставка НДС
        if col_vat_rate:
            actual_row, actual_col = find_merged_cell_for_column(ws, current_row, col_vat_rate)
            safe_set_cell(ws, actual_row, actual_col, vat_rate_str)
        
        # Сумма НДС
        if col_vat_amount:
            actual_row, actual_col = find_merged_cell_for_column(ws, current_row, col_vat_amount)
            safe_set_cell(ws, actual_row, actual_col, float(vat_amount))
        
        # Стоимость с НДС
        if col_total_with_vat:
            actual_row, actual_col = find_merged_cell_for_column(ws, current_row, col_total_with_vat)
            safe_set_cell(ws, actual_row, actual_col, float(amount_with_vat))
        
        current_row += 1
    
    # --- Итоговая строка "Всего к оплате" ---
    # Ищем строку с итогами
    total_row = None
    for row_idx in range(current_row, current_row + 5):
        row_values = [str(ws.cell(row=row_idx, column=col).value or "").lower() 
                     for col in range(1, ws.max_column + 1)]
        if any(keyword in " ".join(row_values) for keyword in ["всего", "итого", "к оплате"]):
            total_row = row_idx
            break
    
    if total_row is None:
        total_row = current_row + 1
    
    # Заполняем итоговые значения (с учетом объединенных ячеек)
    if col_total_no_vat:
        actual_row, actual_col = find_merged_cell_for_column(ws, total_row, col_total_no_vat)
        safe_set_cell(ws, actual_row, actual_col, float(total_without_vat))
    if col_vat_amount:
        actual_row, actual_col = find_merged_cell_for_column(ws, total_row, col_vat_amount)
        safe_set_cell(ws, actual_row, actual_col, float(total_vat))
    if col_total_with_vat:
        actual_row, actual_col = find_merged_cell_for_column(ws, total_row, col_total_with_vat)
        safe_set_cell(ws, actual_row, actual_col, float(total_with_vat))
    
    # Сохраняем файл
    if output_path:
        wb.save(output_path)
        with open(output_path, "rb") as f:
            return f.read()
    else:
        # Возвращаем байты
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()


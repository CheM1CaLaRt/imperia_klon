"""
Модуль для заполнения Универсального передаточного документа (УПД) в Excel
по готовому шаблону Blank-UPD.xlsx
"""
import os
from datetime import date
from decimal import Decimal
from typing import List, Dict, Optional

import openpyxl
from openpyxl import utils


# Путь к шаблону УПД
TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
    "Blank-UPD.xlsx"
)

# Константы с адресами ячеек в шаблоне УПД
# Адреса определены на основе анализа шаблона Blank-UPD.xlsx
CELLS = {
    # Номер и дата документа (поле (1))
    "doc_number": "R2",       # Номер счета-фактуры
    "doc_date": "AA2",        # Дата счета-фактуры
    
    # Продавец (поля (2), (2а), (2б))
    "seller_name": "Y5",       # Продавец (наименование)
    "seller_address": "Y6",   # Адрес продавца
    "seller_inn_kpp": "Y7",   # ИНН/КПП продавца
    "seller_shipper": "Y8",   # Грузоотправитель (опционально)
    "seller_consignee": "Y9", # Грузополучатель (опционально)
    
    # Покупатель (поля (6), (6а), (6б))
    "buyer_name": "Y12",      # Покупатель (наименование)
    "buyer_address": "Y13",   # Адрес покупателя
    "buyer_inn_kpp": "Y14",   # ИНН/КПП покупателя
    
    # Валюта (поле (7))
    "currency": "Y15",        # Валюта
    
    # Таблица товаров - начало первой строки данных
    "table_start_row": 20,    # Строка 20 - первая строка данных товаров
    
    # Колонки таблицы товаров (определены на основе анализа строки 18-20)
    "col_num": 9,             # I - Номер п/п
    "col_name": 12,           # L - Наименование товара
    "col_unit_code": 21,      # U - Код единицы измерения
    "col_unit": 23,           # W - Условное обозначение единицы измерения (национальное)
    "col_price_unit": 30,     # AD - Цена (единицы)
    "col_price": 34,          # AH - Цена (единицы) на единицу измерения
    "col_total_no_vat": 39,   # AM - Стоимость товаров (без НДС)
    "col_vat_rate": 50,       # AX - Ставка НДС
    "col_vat_amount": 54,     # BB - Сумма НДС, начисленная покупателю
    "col_total_with_vat": 61, # BH - Стоимость товаров (с НДС)
    
    # Итоговая строка "Всего к оплате (9)"
    "total_row": 26,          # Строка 26 - итоги
    "total_col_no_vat": 41,   # AO - Сумма без НДС (в формуле =SUM(AM21:AS25))
    "total_col_vat": 56,      # BD - Сумма НДС (в формуле =SUM(BB21:BG25))
    "total_col_with_vat": 62, # BJ - Сумма с НДС (в формуле =SUM(BH21:BN25))
}


def safe_set_cell(worksheet, cell_address: str, value):
    """
    Безопасно устанавливает значение ячейки, обрабатывая объединенные ячейки.
    
    Если ячейка является частью объединенного диапазона, значение записывается
    в верхнюю левую ячейку этого диапазона.
    
    Args:
        worksheet: Рабочий лист openpyxl
        cell_address: Адрес ячейки (например, "BA4")
        value: Значение для установки
    """
    try:
        cell = worksheet[cell_address]
        row = cell.row
        col = cell.column
        
        # Проверяем, является ли ячейка частью объединенного диапазона
        for merged_range in list(worksheet.merged_cells.ranges):
            if (row >= merged_range.min_row and row <= merged_range.max_row and
                col >= merged_range.min_col and col <= merged_range.max_col):
                # Находим верхнюю левую ячейку объединенного диапазона
                top_left_coord = f"{utils.get_column_letter(merged_range.min_col)}{merged_range.min_row}"
                top_left_cell = worksheet[top_left_coord]
                top_left_cell.value = value
                return
        
        # Если ячейка не объединена, устанавливаем значение напрямую
        cell.value = value
    except Exception as e:
        # В случае ошибки пробуем установить значение напрямую
        try:
            worksheet[cell_address].value = value
        except Exception:
            pass


def safe_set_cell_by_coords(worksheet, row: int, col: int, value):
    """
    Безопасно устанавливает значение ячейки по координатам (row, col).
    
    Args:
        worksheet: Рабочий лист openpyxl
        row: Номер строки (начиная с 1)
        col: Номер колонки (начиная с 1)
        value: Значение для установки
    """
    try:
        cell_coord = f"{utils.get_column_letter(col)}{row}"
        safe_set_cell(worksheet, cell_coord, value)
    except Exception:
        try:
            cell = worksheet.cell(row=row, column=col)
            cell.value = value
        except Exception:
            pass


def fill_upd(
    seller_name: str,
    seller_address: str,
    seller_inn_kpp: str,
    buyer_name: str,
    buyer_address: str,
    buyer_inn_kpp: str,
    doc_number: str,
    doc_date: date,
    items: List[Dict],
    currency_name_code: str = "Российский рубль, 643",
    delivery_address: Optional[str] = None,
    output_path: Optional[str] = None,
) -> str:
    """
    Заполняет шаблон УПД данными и сохраняет результат.
    
    Args:
        seller_name: Наименование продавца
        seller_address: Адрес продавца
        seller_inn_kpp: ИНН/КПП продавца (формат: "ИНН / КПП" или просто "ИНН")
        buyer_name: Наименование покупателя
        buyer_address: Адрес покупателя
        buyer_inn_kpp: ИНН/КПП покупателя
        doc_number: Номер документа
        doc_date: Дата документа
        items: Список товаров/услуг, каждый элемент - словарь с ключами:
            - name: наименование
            - unit: единица измерения (условное обозначение)
            - qty: количество
            - price: цена за единицу
            - vat_rate: ставка НДС ("20%", "0%", "без НДС")
        currency_name_code: Валюта (по умолчанию "Российский рубль, 643")
        delivery_address: Адрес доставки (если не указан, используется адрес покупателя)
        output_path: Путь для сохранения файла (если None, создается UPD-filled.xlsx)
    
    Returns:
        str: Путь к сохраненному файлу
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Шаблон УПД не найден: {TEMPLATE_PATH}")
    
    # Загружаем шаблон
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    # Открываем лист "Счет-фактура"
    if "Счет-фактура" in wb.sheetnames:
        ws = wb["Счет-фактура"]
    else:
        # Если лист не найден, используем активный
        ws = wb.active
    
    # --- Заполнение шапки УПД ---
    
    # Номер и дата документа (поле (1))
    safe_set_cell(ws, CELLS["doc_number"], doc_number)
    safe_set_cell(ws, CELLS["doc_date"], doc_date.strftime("%d.%m.%Y"))
    
    # Продавец (поля (2), (2а), (2б))
    safe_set_cell(ws, CELLS["seller_name"], seller_name)
    safe_set_cell(ws, CELLS["seller_address"], seller_address)
    safe_set_cell(ws, CELLS["seller_inn_kpp"], seller_inn_kpp)
    
    # Грузоотправитель (Y8) - адрес продавца
    safe_set_cell(ws, CELLS["seller_shipper"], seller_address)
    
    # Грузополучатель (Y9) - адрес доставки или адрес покупателя
    consignee_address = delivery_address if delivery_address else buyer_address
    safe_set_cell(ws, CELLS["seller_consignee"], consignee_address)
    
    # Покупатель (поля (6), (6а), (6б))
    safe_set_cell(ws, CELLS["buyer_name"], buyer_name)
    safe_set_cell(ws, CELLS["buyer_address"], buyer_address)
    safe_set_cell(ws, CELLS["buyer_inn_kpp"], buyer_inn_kpp)
    
    # Валюта (поле (7))
    safe_set_cell(ws, CELLS["currency"], currency_name_code)
    
    # --- Заполнение таблицы товаров ---
    
    table_start_row = CELLS["table_start_row"]
    current_row = table_start_row
    
    total_without_vat = Decimal("0")
    total_vat = Decimal("0")
    total_with_vat = Decimal("0")
    
    # Словарь для кодов единиц измерения (ОКЕИ)
    # Стандартные коды: 796 - штука, 778 - условная единица, 383 - рубль и т.д.
    # Если единица не найдена, используется код 796 (штука)
    unit_codes = {
        "шт": "796",
        "штука": "796",
        "штуки": "796",
        "усл": "778",
        "услуга": "778",
        "услуги": "778",
        "мес": "мес",
        "месяц": "мес",
        "час": "час",
        "часы": "час",
        "кг": "166",
        "килограмм": "166",
        "т": "168",
        "тонна": "168",
        "м": "006",
        "метр": "006",
        "м2": "055",
        "м²": "055",
        "м3": "113",
        "м³": "113",
    }
    
    # Заполняем строки товаров
    for idx, item in enumerate(items, start=1):
        name = item.get("name", "") or item.get("title", "")  # Поддержка обоих полей
        unit = item.get("unit", "шт") or "шт"  # По умолчанию штука
        qty = float(item.get("qty", 0) or item.get("quantity", 0) or 0)  # Поддержка обоих полей
        price = float(item.get("price", 0) or 0)
        vat_rate_str = item.get("vat_rate", "20%") or "20%"
        
        # Получаем код единицы измерения
        unit_lower = str(unit).lower().strip()
        unit_code = unit_codes.get(unit_lower, "796")  # По умолчанию штука (796)
        
        # Вычисляем суммы
        amount_without_vat = Decimal(str(qty * price))
        
        # Парсим ставку НДС
        if isinstance(vat_rate_str, str):
            if vat_rate_str.endswith("%"):
                vat_percent = Decimal(vat_rate_str.strip("%")) / Decimal("100")
            elif "без" in vat_rate_str.lower() or "ноль" in vat_rate_str.lower():
                vat_percent = Decimal("0")
            else:
                vat_percent = Decimal("0.20")  # По умолчанию 20%
        else:
            vat_percent = Decimal(str(vat_rate_str)) / Decimal("100") if vat_rate_str else Decimal("0.20")
        
        vat_amount = amount_without_vat * vat_percent
        amount_with_vat = amount_without_vat + vat_amount
        
        # Суммируем итоги
        total_without_vat += amount_without_vat
        total_vat += vat_amount
        total_with_vat += amount_with_vat
        
        # Заполняем ячейки строки товара согласно структуре шаблона
        safe_set_cell_by_coords(ws, current_row, CELLS["col_num"], idx)  # I - Номер п/п
        safe_set_cell_by_coords(ws, current_row, CELLS["col_name"], name)  # L - Наименование товара
        safe_set_cell_by_coords(ws, current_row, CELLS["col_unit_code"], unit_code)  # U - Код единицы измерения
        safe_set_cell_by_coords(ws, current_row, CELLS["col_unit"], unit)  # W - Условное обозначение единицы измерения
        safe_set_cell_by_coords(ws, current_row, CELLS["col_price_unit"], price)  # AD - Цена (единицы)
        safe_set_cell_by_coords(ws, current_row, CELLS["col_price"], price)  # AH - Цена (единицы) на единицу измерения
        safe_set_cell_by_coords(ws, current_row, CELLS["col_total_no_vat"], float(amount_without_vat))  # AM - Стоимость без НДС
        safe_set_cell_by_coords(ws, current_row, CELLS["col_vat_rate"], vat_rate_str)  # AX - Ставка НДС
        safe_set_cell_by_coords(ws, current_row, CELLS["col_vat_amount"], float(vat_amount))  # BB - Сумма НДС
        safe_set_cell_by_coords(ws, current_row, CELLS["col_total_with_vat"], float(amount_with_vat))  # BH - Стоимость с НДС
        
        current_row += 1
    
    # --- Заполнение итоговой строки "Всего к оплате (9)" ---
    # Строка 26 содержит итоги в колонках AO (без НДС), BD (НДС), BJ (с НДС)
    
    total_row = CELLS["total_row"]
    
    # Заполняем итоговые значения в правильных колонках
    safe_set_cell_by_coords(ws, total_row, CELLS["total_col_no_vat"], float(total_without_vat))  # AO26 - Сумма без НДС
    safe_set_cell_by_coords(ws, total_row, CELLS["total_col_vat"], float(total_vat))  # BD26 - Сумма НДС
    safe_set_cell_by_coords(ws, total_row, CELLS["total_col_with_vat"], float(total_with_vat))  # BJ26 - Сумма с НДС
    
    # --- Сохранение файла ---
    
    if output_path is None:
        # Создаем имя файла на основе номера документа и даты
        output_path = os.path.join(
            os.path.dirname(TEMPLATE_PATH),
            f"UPD-filled-{doc_number}-{doc_date.strftime('%Y%m%d')}.xlsx"
        )
    
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # Пример использования
    
    # Тестовые данные
    items_data = [
        {
            "name": "Услуги по разработке сайта",
            "unit": "усл",
            "qty": 1,
            "price": 50000.00,
            "vat_rate": "20%",
        },
        {
            "name": "SEO сопровождение",
            "unit": "мес",
            "qty": 1,
            "price": 15000.00,
            "vat_rate": "20%",
        },
        {
            "name": "Консультационные услуги",
            "unit": "час",
            "qty": 5,
            "price": 2000.00,
            "vat_rate": "20%",
        },
    ]
    
    # Заполняем УПД
    output_file = fill_upd(
        seller_name='ООО "Ромашка"',
        seller_address="197000, г. Санкт-Петербург, ул. Примерная, д. 1",
        seller_inn_kpp="7700000000 / 770001001",
        buyer_name='ООО "Покупатель"',
        buyer_address="101000, г. Москва, ул. Клиента, д. 2",
        buyer_inn_kpp="7711111111 / 771101001",
        doc_number="15",
        doc_date=date.today(),
        items=items_data,
    )
    
    print(f"УПД успешно создан: {output_file}")

